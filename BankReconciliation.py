from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from csv import reader
from openpyxl import load_workbook
import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side


#global variables
bank_statement_path = None
general_ledger_path = None

#solution found at https://stackoverflow.com/questions/44033894/removing-common-values-from-two-lists-in-python/44033987
def remove_values_from_list(the_list, val):
   return [value for value in the_list if value != val] 

def sheet_setup(sheet):
    
    sheet['A1'] = 'Bank Statement'
    sheet['F1'] = 'General Ledger'
    sheet['A2'] = 'Date'
    sheet['B2'] = 'Source Num'
    sheet['C2'] = 'Comment'
    sheet['D2'] = 'Debit'
    sheet['E2'] = 'Credit'
    sheet['F2'] = 'Date'
    sheet['G2'] = 'Source Num'
    sheet['H2'] = 'Comment'
    sheet['I2'] = 'Debit'
    sheet['J2'] = 'Credit'

    sheet['A1'].font = Font(size=14, underline="single", bold=True)
    sheet['F1'].font = Font(size=14, underline="single", bold=True)

    

    row = sheet['A2':'J2']
    row = row[0]
    
    for cell in row:
        cell.border = Border(bottom=Side(border_style="thin"))
        cell.font = Font(bold=True)

    sheet['E2'].border = Border(right=Side(border_style="thin"), bottom=Side(border_style='thin'))
    
    


def populate(sheet, bank_or_ledger, list):
    if bank_or_ledger == 'bank_statement':
        cells = ['A', 'B', 'C', 'D', 'E']
    else:
        cells = ['F', 'G', 'H', 'I', 'J']
    entry_order = ['date', 'source_num', 'comment', 'debit', 'credit']

    for i in range(3, len(list) + 3):
        for j in range(len(cells)):
            cell_index = cells[j] + str(i)
            sheet[cell_index] = list[i-3][entry_order[j]]
        e_cell = 'E' + str(i)
        sheet[e_cell].border = Border(right=Side(border_style='thin'))

#solution by velis at https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
def resize_sheet_columns(sheet):
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value + 2

def reconcile():
    if (bank_statement_path == None or general_ledger_path == None):
        pass
    else:   
        CSV = processCSV()
        excel = processExcel()
        ascending_dates(CSV)
        ascending_dates(excel)
        
        entry_lists = {}
        
        entry_lists["matching_cheques"] = {}
        entry_lists["matching_cheques"]["bank_statements"] = []
        entry_lists["matching_cheques"]["general_ledger"] = []
        entry_lists["matching_cheques"]["total_credit"] = [0, 0]
        entry_lists["matching_cheques"]["total_debit"] = [0, 0]

        entry_lists["canada_helps"] = {}
        entry_lists["canada_helps"]["bank_statements"] = []
        entry_lists["canada_helps"]["general_ledger"] = []
        entry_lists["canada_helps"]["total_debit"] = [0, 0]

        entry_lists["paypal"] = {}
        entry_lists["paypal"]["bank_statements"] = []
        entry_lists["paypal"]["general_ledger"] = []
        entry_lists["paypal"]["total_credit"] = [0, 0]
        entry_lists["paypal"]["total_debit"] = [0, 0]

        entry_lists["etransfer"] = {}
        entry_lists["etransfer"]["bank_statements"] = []
        entry_lists["etransfer"]["general_ledger"] = []
        entry_lists["etransfer"]["total_credit"] = [0, 0]
        entry_lists["etransfer"]["total_debit"] = [0, 0]
                
        for i in range(len(CSV)):
            if 'canada help' in CSV[i]["comment"].lower():
                entry_lists["canada_helps"]["bank_statements"].append(CSV[i])
                entry_lists["canada_helps"]["total_debit"][0] += float(CSV[i]["debit"])
                CSV[i] = 0

            elif 'email money tran' in CSV[i]["comment"].lower():
                entry_lists["etransfer"]["bank_statements"].append(CSV[i])
                entry_lists["etransfer"]["total_debit"][0] += float(CSV[i]["debit"])
                entry_lists["etransfer"]["total_credit"][0] += float(CSV[i]["credit"])
                CSV[i] = 0

            elif 'paypal' in CSV[i]["comment"].lower() or 'pay pal' in CSV[i]["comment"].lower():
                entry_lists["paypal"]["bank_statements"].append(CSV[i])
                entry_lists["paypal"]["total_debit"][0] += float(CSV[i]["debit"])
                CSV[i] = 0

            elif CSV[i]["source_num"] != "":
                for j in range(len(excel)):
                    if excel[j] != 0:
                        if CSV[i]["source_num"] == excel[j]["source_num"]:
                            
                            entry_lists["matching_cheques"]["bank_statements"].append(CSV[i])
                            entry_lists["matching_cheques"]["general_ledger"].append(excel[j])
                            entry_lists["matching_cheques"]["total_debit"][0] += float(CSV[i]["debit"])
                            entry_lists["matching_cheques"]["total_credit"][0] += float(CSV[i]["credit"])

                            entry_lists["matching_cheques"]["total_credit"][1] += float(excel[j]["credit"])
                            entry_lists["matching_cheques"]["total_debit"][1] += float(excel[j]["debit"])

                            CSV[i] = 0
                            excel[j] = 0
                            break
        
        for i in range(len(excel)):
            if excel[i] != 0:
                if 'canada' in excel[i]["source_num"].lower():
                    entry_lists["canada_helps"]["general_ledger"].append(excel[i])
                    entry_lists["canada_helps"]["total_debit"][1] += float(excel[i]["debit"])
                    excel[i] = 0

                elif 'etransfer' in excel[i]["comment"].lower() or 'e transfer' in excel[i]["comment"].lower():
                    entry_lists["etransfer"]["general_ledger"].append(excel[i])
                    entry_lists["etransfer"]["total_debit"][1] += float(excel[i]["debit"])
                    entry_lists["etransfer"]["total_credit"][1] += float(excel[i]["credit"])
                    excel[i] = 0

                elif 'paypal' in excel[i]["comment"].lower() or 'pay pal' in excel[i]["comment"].lower():
                    entry_lists["paypal"]["general_ledger"].append(excel[i])
                    entry_lists["paypal"]["total_credit"][1] += float(excel[i]["credit"])
                    entry_lists["paypal"]["total_debit"][1] += float(excel[i]["debit"])
                    excel[i] = 0

        CSV = remove_values_from_list(CSV, 0)
        excel = remove_values_from_list(excel, 0)

        work_book = Workbook()
        wb_filename = 'bank_rec_' + str(datetime.date.today()) + '.xlsx'
        print(wb_filename)

        matched_cheques = work_book.active
        matched_cheques.title = 'Matched Cheques'

        canada_helps = work_book.create_sheet(title="Canada Helps")
        paypal = work_book.create_sheet(title="Paypal")
        etransfer = work_book.create_sheet(title="Etransfer")
        unmatched_entries = work_book.create_sheet(title="Unmatched Sheets")

        sheets = []
        sheets.append(matched_cheques)
        sheets.append(canada_helps)
        sheets.append(paypal)
        sheets.append(etransfer)
        sheets.append(unmatched_entries)

        for sheet in sheets:
            sheet_setup(sheet)

        populate(matched_cheques, 'bank_statement', entry_lists["matching_cheques"]["bank_statements"])
        populate(matched_cheques, 'general_ledger', entry_lists["matching_cheques"]["general_ledger"])

        populate(canada_helps, 'bank_statement', entry_lists["canada_helps"]["bank_statements"])
        populate(canada_helps, 'general_ledger', entry_lists["canada_helps"]["general_ledger"])

        populate(paypal, 'bank_statement', entry_lists["paypal"]["bank_statements"])
        populate(paypal, 'general_ledger', entry_lists["paypal"]["general_ledger"])

        populate(etransfer, 'bank_statement', entry_lists["etransfer"]["bank_statements"])
        populate(etransfer, 'general_ledger', entry_lists["etransfer"]["general_ledger"])

        populate(unmatched_entries, 'bank_statement', CSV)
        populate(unmatched_entries, 'general_ledger', excel)

        for sheet in sheets:
            resize_sheet_columns(sheet)

        
        work_book.save(filename = wb_filename)


        #testing print statements
        '''
        print("Matched cheques:\n")
        for pair in entry_lists["matching_cheques"]["cheques"]:
            print(f"bank statement: {pair[0]} \ngeneral ledger: {pair[1]}\n")
        

        print(f"Matched Cheques: \nBank Deposits: credit: ${entry_lists['matching_cheques']['total_credit'][0]}", \
            f"debit: ${entry_lists['matching_cheques']['total_debit'][0]}", \
            f"\nGeneral Ledger: credit: ${entry_lists['matching_cheques']['total_credit'][1]}", \
            f"debit: ${entry_lists['matching_cheques']['total_debit'][1]}\n")

        print(f"Canada Helps:\nBank Statement Debits: ${entry_lists['canada_helps']['total_debit'][0]}", \
            f"General Ledger Debits: ${entry_lists['canada_helps']['total_debit'][1]}\n")

        print(f"Paypal: \nBank Deposits: credit: ${entry_lists['paypal']['total_credit'][0]}", \
            f"debit: ${entry_lists['paypal']['total_debit'][0]}", \
            f"\nGeneral Ledger: credit: ${entry_lists['paypal']['total_credit'][1]}", \
            f"debit: ${entry_lists['paypal']['total_debit'][1]}\n")

        print(f"E transfer: \nBank Deposits: credit: ${entry_lists['etransfer']['total_credit'][0]}", \
            f"debit: ${entry_lists['etransfer']['total_debit'][0]}", \
            f"\nGeneral Ledger: credit: ${entry_lists['etransfer']['total_credit'][1]}", \
            f"debit: ${entry_lists['etransfer']['total_debit'][1]}\n")
        '''


#found at https://stackoverflow.com/questions/8270092/remove-all-whitespace-in-a-string
def removeExtraSpaces(string):
    return(" ".join(string.split()))

def ascending_dates(list):
    first_date = list[0]["date"]
    last_date = list[-1]["date"]

    first_day = first_date.split("-")[2]
    last_day = last_date.split("-")[2]

    if int(first_day) > int(last_day):
        list.reverse()



def standardize_date_string(string):
    split_string = string.split('-')
    new_string = ''
    year = split_string[2]
    new_string += '20' + year + '-'
    orig_month = split_string[1]
    if len(orig_month) != 3:
        orig_month = orig_month[:3]

    months = {"Jan":"01","Feb":"02","Mar":"03","Apr":"04","May":"05","Jun":"06","Jul":"07","Aug":"08","Sep":"09","Oct":"10","Nov":"11","Dec":"12"}
    new_month = months[orig_month]

    new_string += new_month + '-'

    new_string += split_string[0]

    return new_string

def processCSV():
    formattedCSV = []
    with open(bank_statement_path, newline='') as csvfile:
        csv_reader = reader(csvfile)
        for row in csv_reader:
            entry = {}
            entry["date"] = standardize_date_string(row[1])
            entry["comment"] = removeExtraSpaces(row[2])
            entry["source_num"] = str(row[3]).strip()
            entry["credit"] = str(row[4]).strip()
            if entry["credit"] == '':
                entry["credit"] = '0'
            entry["debit"] = str(row[5]).strip()
            if entry["debit"] == '':
                entry["debit"] = '0'
            

            formattedCSV.append(entry)
    print(formattedCSV[0]["date"])
    print(type(formattedCSV[0]["date"]))
    return formattedCSV
    
    
def processExcel():
    workbook = load_workbook(filename=general_ledger_path, read_only=True)
    sheet = workbook['Sheet1']
    rows = list(sheet.rows)
    rows = rows[5:-2]
    formattedExcel = []
    for row in rows:
        data = []
        for cell in row:
            data.append(cell.value)
        entry = {}

        entry["date"] = str(data[2])[0:10]
        entry["comment"] = str(data[3]).strip()
        entry["source_num"] = str(data[4]).strip()
        entry["debit"] = str(data[6]).strip()
        entry["credit"] = str(data[7]).strip()

        formattedExcel.append(entry)
    return formattedExcel

        

def select_bank_file():
    global bank_statement_path
    bank_statement_path = filedialog.askopenfilename(initialdir = "/",title = "Select Bank File",filetypes = ( ("csv files","*.csv"), ))
    bank_statement_name.set(bank_statement_path.split('/')[-1])

def select_sage_file():
    global general_ledger_path
    general_ledger_path = filedialog.askopenfilename(initialdir = "/",title = "Select Sage File",filetypes = ( ("xlsx files","*.xlsx"), ))
    general_ledger_name.set(general_ledger_path.split('/')[-1])
    



root = Tk()
root.title("Bank Reconciliation")

mainframe = ttk.Frame(root, padding="3 3 12 12")
mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

bank_statement_name = StringVar()
bank_statement_name.set('None')

general_ledger_name = StringVar()
general_ledger_name.set('None')

ttk.Label(mainframe, text='Select Bank Statement:').grid(column=0, row=0, sticky=W, padx=(50, 15), pady=5)
ttk.Button(mainframe, text='Choose .csv File', command=select_bank_file).grid(column=1, row=0, padx=(15,50), pady=5)
ttk.Label(mainframe, text='Selected File:').grid(column=0, row=1, padx=(50, 15), pady=5, sticky=W)
ttk.Label(mainframe, textvariable=bank_statement_name).grid(column=1, row=1, padx = 15, pady=5, sticky=W)

ttk.Label(mainframe, text='Select General Ledger:').grid(column=2, row=0, sticky=W, padx=(50, 15), pady=5)
ttk.Button(mainframe, text='Choose .xlsx File', command=select_sage_file).grid(column=3, row=0, padx=(15, 50), pady=5)
ttk.Label(mainframe, text='Selected File:').grid(column=2, row=1, padx=(50,15), pady=5, sticky=W)
ttk.Label(mainframe, textvariable=general_ledger_name).grid(column=3, row=1, padx=(15,50), pady=5, sticky=W)

ttk.Button(mainframe, text='Reconcile', command=reconcile).grid(column=4, row=0, rowspan=2, padx=(50,15), pady=15)


root.mainloop()